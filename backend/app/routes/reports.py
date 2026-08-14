from fastapi import APIRouter, Depends, HTTPException, Response
from sqlalchemy.orm import Session, joinedload, selectinload
from sqlalchemy import func
from typing import Optional
from datetime import date
from dateutil.relativedelta import relativedelta
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pytz
from datetime import datetime

from ..database import get_db
from .. import models
from ..auth import get_current_user, get_query
from ..permissions import has_permission
from ..services.low_stock import get_low_stock_items

# 👇 IMPORT LOGIKA DARI ACCOUNTING (SINGLE SOURCE OF TRUTH) 👇
from .accounting import get_income_statement 

router = APIRouter()

WITA = pytz.timezone("Asia/Makassar")
REAL_PURCHASE_STATUSES = ["unpaid", "partial", "paid"]

def get_local_date():
    return datetime.now(WITA).date()


def _require_financial_report_access(db: Session, user: models.User):
    if not has_permission(db, user, "report.financial", "view"):
        raise HTTPException(403, "Akses laporan modal/HPP/laba hanya untuk admin")

# ─── 1. DASHBOARD KPI (UTAMA) ────────────────────────────────────────────────
@router.get("/dashboard")
def get_dashboard_data(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    today_local = get_local_date() 

    total_sales_today = get_query(db, models.Sale, current_user).filter(
        models.Sale.date == today_local,
        models.Sale.status != "cancelled"
    ).with_entities(func.sum(models.Sale.total)).scalar() or 0

    purchase_q = db.query(models.Purchase).filter(
        models.Purchase.is_branch_request == False,
        models.Purchase.date == today_local,
        models.Purchase.status.in_(REAL_PURCHASE_STATUSES)
    )
    if current_user.active_branch_id is not None:
        purchase_q = purchase_q.filter(models.Purchase.branch_id == current_user.active_branch_id)
    total_purchases_today = purchase_q.with_entities(func.sum(models.Purchase.total)).scalar() or 0

    total_tx_today = get_query(db, models.Sale, current_user).filter(
        models.Sale.date == today_local,
        models.Sale.status != "cancelled"
    ).count()

    # 👇 BARU: Ambil Laba Bersih Bulan Ini dari Modul Akuntansi
    net_profit_month = 0
    if has_permission(db, current_user, "report.financial", "view"):
        start_month = today_local.replace(day=1)
        acc_report = get_income_statement(start_date=start_month, end_date=today_local, db=db, current_user=current_user)
        net_profit_month = acc_report.get("net_profit", 0)

    # Satu sumber data dengan tab Persediaan > Stok Menipis agar jumlah dan
    # rincian barang yang terlihat pengguna selalu konsisten.
    low_stock_items = get_low_stock_items(db, current_user.active_branch_id)

    # Top 5 Produk Terlaris Bulan Ini (Filter Cabang)
    top_items_raw = get_query(db, models.Sale, current_user).join(
        models.SaleItem
    ).join(
        models.Item
    ).with_entities(
        models.Item.name,
        func.sum(models.SaleItem.qty).label("total_qty"),
        func.sum(models.SaleItem.total).label("total_amount")
    ).filter(
        models.Sale.date >= today_local.replace(day=1),
        models.Sale.date < today_local.replace(day=1) + relativedelta(months=1),
        models.Sale.status != "cancelled"
    ).group_by(models.Item.id).order_by(func.sum(models.SaleItem.qty).desc()).limit(5).all()
    
    top_items = [{"name": r[0], "qty": r[1], "amount": r[2]} for r in top_items_raw]

    # Total Deposit
    total_cust_deposit = db.query(func.sum(models.Customer.deposit_balance)).scalar() or 0
    total_supp_deposit = db.query(func.sum(models.Supplier.deposit_balance)).scalar() or 0

    # 5 Penjualan Terbaru (Filter Cabang)
    recent_sales = [{
        "number": s.number, 
        "date": str(s.date),
        "customer": s.customer.name if s.customer else "Umum",
        "total": s.total, 
        "status": s.status
    } for s in get_query(db, models.Sale, current_user).options(
        joinedload(models.Sale.customer)
    ).order_by(models.Sale.id.desc()).limit(5).all()]

    # Grafik 6 Bulan Terakhir (Filter Cabang)
    monthly = []
    for i in range(5, -1, -1):
        d = today_local - relativedelta(months=i)
        month_start = d.replace(day=1)
        next_month = month_start + relativedelta(months=1)
        # Range tanggal (ramah index ix_sales_branch_date) — bukan cast().like() yang full-scan.
        amt = get_query(db, models.Sale, current_user).filter(
            models.Sale.date >= month_start,
            models.Sale.date < next_month,
            models.Sale.status != "cancelled"
        ).with_entities(func.sum(models.Sale.total)).scalar() or 0
        monthly.append({"month": d.strftime("%b %Y"), "amount": float(amt)})

    return {
        "total_sales_today": float(total_sales_today),
        "total_purchases_today": float(total_purchases_today),
        "total_transactions_today": int(total_tx_today),
        "net_profit_monthly": float(net_profit_month), # 👈 Data Net Profit Akurat
        "low_stock_count": len(low_stock_items),
        "low_stock_items": low_stock_items,
        "total_customer_deposit": float(total_cust_deposit),
        "total_supplier_deposit": float(total_supp_deposit),
        "top_items": top_items,
        "recent_sales": recent_sales,
        "monthly_sales": monthly
    }

# ─── 2. LABA RUGI (SINKRON DENGAN AKUNTANSI) ─────────────────────────────────
@router.get("/profit-loss")
def profit_loss(
    start_date: Optional[date] = None, 
    end_date: Optional[date] = None, 
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(get_current_user)
):
    # 👇 REVISI TOTAL: Tanya langsung ke modul akuntansi
    _require_financial_report_access(db, current_user)
    acc_data = get_income_statement(start_date=start_date, end_date=end_date, db=db, current_user=current_user)
    period_start = start_date or get_local_date().replace(day=1)
    period_end = end_date or get_local_date()
    purchase_q = db.query(models.Purchase).filter(
        models.Purchase.is_branch_request == False,
        models.Purchase.date >= period_start,
        models.Purchase.date <= period_end,
        models.Purchase.status.in_(REAL_PURCHASE_STATUSES)
    )
    if current_user.active_branch_id is not None:
        purchase_q = purchase_q.filter(models.Purchase.branch_id == current_user.active_branch_id)
    total_purchases = purchase_q.with_entities(func.sum(models.Purchase.total)).scalar() or 0
    
    return {
        # Format Baru (Lengkap)
        "total_revenue": acc_data.get("total_revenue", 0),
        "total_cogs": acc_data.get("total_cogs", 0),
        "gross_profit": acc_data.get("gross_profit", 0),
        "total_operating_expense": acc_data.get("total_operating_expense", 0),
        "total_other_expense": acc_data.get("total_other_expense", 0),
        "net_profit": acc_data.get("net_profit", 0),
        "period": acc_data.get("period"),
        
        # Format Lama (Backward Compatibility untuk mencegah Frontend Error)
        "total_sales": acc_data.get("total_revenue", 0),
        "total_purchases": float(total_purchases),
        "hpp": acc_data.get("total_cogs", 0), 
        "expenses": acc_data.get("total_operating_expense", 0) + acc_data.get("total_other_expense", 0),
        "transaction_count": get_query(db, models.Sale, current_user).filter(
            models.Sale.date >= period_start,
            models.Sale.date <= period_end,
            models.Sale.status != "cancelled"
        ).count()
    }

# ─── 3. HUTANG & PIUTANG ──────────────────────────────────────────────────────
@router.get("/receivables")
def receivables(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    sales = get_query(db, models.Sale, current_user).options(
        joinedload(models.Sale.customer)
    ).filter(
        models.Sale.status.in_(["unpaid", "partial"])
    ).all()
    return [{
        "id": s.id,
        "date": s.date,
        "number": s.number,
        "customer": s.customer.name if s.customer else "Umum",
        "total": s.total,
        "paid": s.paid,
        "remaining": s.total - s.paid
    } for s in sales]

@router.get("/payables")
def payables(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    q = db.query(models.Purchase).filter(
        models.Purchase.is_branch_request == False,
        models.Purchase.status.in_(["unpaid", "partial"])
    )
    if current_user.active_branch_id is not None:
        q = q.filter(models.Purchase.branch_id == current_user.active_branch_id)
    purchases = q.options(joinedload(models.Purchase.supplier)).all()

    return [{
        "id": p.id,
        "date": p.date,
        "number": p.number, 
        "supplier": p.supplier.name if p.supplier else "-", 
        "total": p.total,
        "paid": p.paid,
        "remaining": p.total - p.paid
    } for p in purchases]

# ─── 4. DETAILED SALES REPORT ────────────────────────────────────────────────
@router.get("/sales-detailed")
def get_sales_detailed(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    today_local = get_local_date()
    if not start_date: start_date = today_local.replace(day=1)
    if not end_date: end_date = today_local

    sales = get_query(db, models.Sale, current_user).options(
        joinedload(models.Sale.customer),
        selectinload(models.Sale.items).joinedload(models.SaleItem.item),
    ).filter(
        models.Sale.date >= start_date,
        models.Sale.date <= end_date,
        models.Sale.status != "cancelled"
    ).all()

    results = []
    for s in sales:
        items = []
        for it in s.items:
            items.append({
                "item_name": it.item.name if it.item else "Unknown",
                "qty": it.qty,
                "price": it.sell_price,
                "discount": it.discount,
                "total": (it.sell_price * (1 - it.discount/100)) * it.qty
            })
        
        results.append({
            "id": s.id,
            "number": s.number,
            "date": str(s.date),
            "customer": s.customer.name if s.customer else "Umum",
            "total": s.total,
            "paid": s.paid,
            "payment_method": s.payment_method,
            "items": items
        })
    
    return results

# ─── 5. INVENTORY VALUATION ──────────────────────────────────────────────────
@router.get("/inventory-valuation")
def get_inventory_valuation(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    _require_financial_report_access(db, current_user)
    # Calculate stock value based on buy_price
    # If branch active, filter by warehouse stocks
    if current_user.active_branch_id:
        gudang_ids = [g.id for g in db.query(models.Warehouse).filter(models.Warehouse.branch_id == current_user.active_branch_id).all()]
        stock_data = db.query(
            models.Item.name,
            models.Item.buy_price,
            models.Item.sell_price,
            func.sum(models.WarehouseStock.stock).label("total_stock")
        ).join(models.WarehouseStock).filter(
            models.WarehouseStock.warehouse_id.in_(gudang_ids),
            models.Item.is_active == True
        ).group_by(models.Item.id).all()
    else:
        stock_data = db.query(
            models.Item.name,
            models.Item.buy_price,
            models.Item.sell_price,
            models.Item.stock.label("total_stock")
        ).filter(models.Item.is_active == True).all()

    results = []
    total_value_buy = 0
    total_value_sell = 0
    
    for row in stock_data:
        val_buy = (row.total_stock or 0) * (row.buy_price or 0)
        val_sell = (row.total_stock or 0) * (row.sell_price or 0)
        total_value_buy += val_buy
        total_value_sell += val_sell
        results.append({
            "name": row.name,
            "stock": row.total_stock or 0,
            "buy_price": row.buy_price or 0,
            "sell_price": row.sell_price or 0,
            "value_at_buy": val_buy,
            "value_at_sell": val_sell
        })
    
    return {
        "items": results,
        "total_inventory_value_buy": total_value_buy,
        "total_inventory_value_sell": total_value_sell
    }

# ─── 6. IMPROVED EXPORT EXCEL ────────────────────────────────────────────────
@router.get("/export/full-report")
def export_full_report(
    start_date: Optional[date] = None, 
    end_date: Optional[date] = None, 
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(get_current_user)
):
    _require_financial_report_access(db, current_user)
    today_local = get_local_date()
    if not start_date: start_date = today_local.replace(day=1)
    if not end_date: end_date = today_local

    wb = openpyxl.Workbook()
    
    # 1. Sheet Penjualan
    ws1 = wb.active
    ws1.title = "Penjualan"
    headers = ["No. Faktur", "Tanggal", "Pelanggan", "Total", "Dibayar", "Metode", "Status"]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    sales = get_query(db, models.Sale, current_user).filter(
        models.Sale.date >= start_date, 
        models.Sale.date <= end_date,
        models.Sale.status != "cancelled" 
    ).all()
    
    for s in sales:
        ws1.append([s.number, str(s.date), s.customer.name if s.customer else "Umum", s.total, s.paid, s.payment_method, s.status])

    # 2. Sheet Detail Item Terjual
    ws2 = wb.create_sheet("Detail Item Terjual")
    ws2.append(["No. Faktur", "Barang", "Qty", "Harga", "Diskon %", "Subtotal"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for s in sales:
        for it in s.items:
            sub = (it.sell_price * (1 - it.discount/100)) * it.qty
            ws2.append([s.number, it.item.name if it.item else "-", it.qty, it.sell_price, it.discount, sub])

    # 3. Sheet Laba Rugi Ringkas
    ws3 = wb.create_sheet("Laba Rugi")
    acc_data = get_income_statement(start_date=start_date, end_date=end_date, db=db, current_user=current_user)
    ws3.append(["Keterangan", "Nilai"])
    ws3.append(["Pendapatan Penjualan (Net)", acc_data.get("total_revenue", 0)])
    ws3.append(["HPP (Harga Pokok Penjualan)", acc_data.get("total_cogs", 0)])
    ws3.append(["Laba Kotor", acc_data.get("gross_profit", 0)])
    ws3.append(["Biaya Operasional", acc_data.get("total_operating_expense", 0)])
    ws3.append(["Biaya Lain-lain", acc_data.get("total_other_expense", 0)])
    ws3.append(["Laba Bersih", acc_data.get("net_profit", 0)])
    for cell in ws3[1]: cell.font = Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Laporan_Lengkap_{start_date}_sd_{end_date}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ─── 7. TOP ITEMS ────────────────────────────────────────────────────────────
@router.get("/top-items")
def get_top_items(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    today_local = get_local_date()
    if not start_date: start_date = today_local.replace(day=1)
    if not end_date: end_date = today_local

    top_items_raw = get_query(db, models.Sale, current_user).join(
        models.SaleItem
    ).join(
        models.Item
    ).with_entities(
        models.Item.name,
        func.sum(models.SaleItem.qty).label("total_qty"),
        func.sum(models.SaleItem.total).label("total_amount")
    ).filter(
        models.Sale.date >= start_date,
        models.Sale.date <= end_date,
        models.Sale.status != "cancelled"
    ).group_by(models.Item.id).order_by(func.sum(models.SaleItem.qty).desc()).limit(limit).all()
    
    return [{"name": r[0], "qty": r[1], "revenue": r[2]} for r in top_items_raw]

# ─── 8. DEPOSIT & RETUR BALANCE ─────────────────────────────────────────────
@router.get("/deposits/customers")
def get_customer_deposits(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    _require_financial_report_access(db, current_user)
    customers = db.query(models.Customer).filter(models.Customer.deposit_balance > 0).all()
    return [{
        "id": c.id,
        "code": c.code,
        "name": c.name,
        "deposit_balance": c.deposit_balance
    } for c in customers]

@router.get("/deposits/suppliers")
def get_supplier_deposits(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    _require_financial_report_access(db, current_user)
    suppliers = db.query(models.Supplier).filter(models.Supplier.deposit_balance > 0).all()
    return [{
        "id": s.id,
        "code": s.code,
        "name": s.name,
        "deposit_balance": s.deposit_balance
    } for s in suppliers]

# ─── 4. EXPORT EXCEL SALES ───────────────────────────────────────────────────
@router.get("/export/sales")
def export_sales(
    start_date: Optional[date] = None, 
    end_date: Optional[date] = None, 
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(get_current_user)
):
    today_local = get_local_date()
    if not start_date: start_date = today_local.replace(day=1)
    if not end_date: end_date = today_local

    sales = get_query(db, models.Sale, current_user).filter(
        models.Sale.date >= start_date, 
        models.Sale.date <= end_date,
        models.Sale.status != "cancelled" 
    ).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["No. Faktur", "Tanggal", "Cabang ID", "Pelanggan", "Total", "Status"])
    
    for s in sales:
        ws.append([s.number, str(s.date), s.branch_id or "Pusat", s.customer.name if s.customer else "Umum", s.total, s.status])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    cbg_name = f"Cabang_{current_user.active_branch_id}" if current_user.active_branch_id else "Semua_Cabang"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Laporan_Penjualan_{cbg_name}_{start_date}_to_{end_date}.xlsx"}
    )
