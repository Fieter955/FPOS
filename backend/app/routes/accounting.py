"""
iPos 5.0 — Akuntansi Lengkap (Enterprise Multi-Branch Edition)
Fitur:
  - Chart of Accounts (Daftar Akun)

  - Jurnal Umum (double-entry) dengan stempel cabang
  - Buku Besar per akun (terisolasi per cabang)
  - Neraca Saldo (Trial Balance) per cabang
  - Neraca (Balance Sheet) per cabang
  - Laporan Laba Rugi (Income Statement) per cabang
  - Kas Masuk / Kas Keluar dengan nomor seri khusus cabang
  - Auto-journal dari transaksi penjualan & pembelian
"""

from fastapi import APIRouter, Depends, HTTPException, Response
import pytz
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, List
from datetime import date, datetime, timedelta # 👈 timedelta ditambahkan di sini
from pydantic import BaseModel
from calendar import monthrange

from ..database import get_db
from .. import models
from ..auth import get_current_user, require_admin, write_audit, get_query
from .. import schemas

router = APIRouter()
WITA = pytz.timezone("Asia/Makassar")

def get_local_date():
    return datetime.now(WITA).date()


def resolve_active_branch_id(db: Session, current_user: models.User) -> int:
    branch_id = current_user.active_branch_id
    if branch_id:
        exists = db.query(models.Branch.id).filter(models.Branch.id == branch_id).first()
        if exists:
            return branch_id

    first_branch = db.query(models.Branch).order_by(models.Branch.id).first()
    if not first_branch:
        raise HTTPException(400, "Sistem belum memiliki cabang sama sekali.")
    return first_branch.id


def get_books_locked_until(db: Session, branch_id: Optional[int]) -> Optional[date]:
    q = db.query(models.BookClose).filter(models.BookClose.is_active == True)
    if branch_id is not None:
        q = q.filter(models.BookClose.branch_id == branch_id)
    else:
        q = q.filter(models.BookClose.branch_id == None)
    obj = q.order_by(models.BookClose.period_end.desc()).first()
    return obj.period_end if obj else None


def assert_books_open(
    db: Session,
    branch_id: Optional[int],
    trx_date: date,
    action_label: str = "Transaksi",
):
    locked_until = get_books_locked_until(db, branch_id)
    if locked_until and trx_date <= locked_until:
        raise HTTPException(
            status_code=400,
            detail=f"DITOLAK: {action_label} tanggal {trx_date} berada di periode tutup buku (sampai {locked_until}).",
        )


class BookCloseCreate(BaseModel):
    close_type: str = "month"  # month | year
    end_date: date
    notes: Optional[str] = None


def get_branch_inventory_snapshot(db: Session, branch_id: int) -> dict:
    warehouse_ids = [
        wid
        for (wid,) in db.query(models.Warehouse.id)
        .filter(models.Warehouse.branch_id == branch_id)
        .all()
    ]

    if warehouse_ids:
        inventory_value = (
            db.query(
                func.sum(
                    models.WarehouseStock.stock * func.coalesce(models.Item.buy_price, 0)
                )
            )
            .join(models.Item, models.Item.id == models.WarehouseStock.item_id)
            .filter(
                models.WarehouseStock.warehouse_id.in_(warehouse_ids),
                models.WarehouseStock.stock > 0,
            )
            .scalar()
            or 0.0
        )
        item_count = (
            db.query(func.count(func.distinct(models.WarehouseStock.item_id)))
            .join(models.Item, models.Item.id == models.WarehouseStock.item_id)
            .filter(
                models.WarehouseStock.warehouse_id.in_(warehouse_ids),
                models.WarehouseStock.stock > 0,
            )
            .scalar()
            or 0
        )
    else:
        inventory_value = (
            db.query(
                func.sum(models.Item.stock * func.coalesce(models.Item.buy_price, 0))
            )
            .filter(models.Item.stock > 0)
            .scalar()
            or 0.0
        )
        item_count = (
            db.query(func.count(models.Item.id))
            .filter(models.Item.stock > 0)
            .scalar()
            or 0
        )

    return {
        "branch_id": branch_id,
        "warehouse_count": len(warehouse_ids),
        "item_count": int(item_count or 0),
        "inventory_value": float(inventory_value or 0.0),
    }

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def next_journal_number(db: Session, branch_id: int) -> str:
    """Generate nomor jurnal unik per cabang untuk menghindari UNIQUE constraint error"""
    today_str = datetime.now(WITA).strftime('%Y%m%d')
    b_id = branch_id or 0
    prefix = f"JU-C{b_id}-{today_str}"
    
    last = db.query(models.Journal).filter(
        models.Journal.number.like(f"{prefix}%")
    ).order_by(models.Journal.id.desc()).with_for_update().first() 
    
    seq = int(last.number[-4:]) + 1 if last else 1
    return f"{prefix}{seq:04d}"


def create_auto_journal(
    db: Session,
    date_val: date,
    number_ref: str,
    description: str,
    entries: list,
    user_id: int,
    branch_id: int,
    allow_closed_period: bool = False,
):
    """Helper untuk membuat jurnal otomatis dari modul lain (Sales/Purchases/dll).

    allow_closed_period=True hanya untuk RESTATEMENT (mis. koreksi HPP swap batch
    Fase 3) yang sengaja memposting jurnal bertanggal di periode yang sudah Tutup
    Buku — karena laporan dihitung live dari baris jurnal, neraca/laba-rugi periode
    lama otomatis ikut menyesuaikan. Jalur ini wajib admin-only + tercatat di audit.
    """
    if not allow_closed_period:
        assert_books_open(db, branch_id, date_val, "Jurnal otomatis")

    # 1. Validasi Keseimbangan (Balance)
    total_debit = sum(e["debit"] for e in entries)
    total_credit = sum(e["credit"] for e in entries)
    if abs(total_debit - total_credit) > 0.01:
        raise ValueError(f"Jurnal tidak balance! Debit: {total_debit}, Kredit: {total_credit}")

    # 👇 2. PENJAGA GERBANG ABSOLUT: Cek apakah akun benar-benar ada di DB 👇
    for entry in entries:
        if entry["debit"] == 0 and entry["credit"] == 0:
            continue
        # Pastikan akun ada
        cek_akun = db.query(models.Account).filter(models.Account.code == entry["code"]).first()
        if not cek_akun:
            raise ValueError(f"FATAL ERROR: Akun dengan kode '{entry['code']}' tidak ditemukan di Chart of Accounts! Transaksi dibatalkan untuk mencegah jurnal tidak seimbang.")

    # 3. Buat Header Jurnal (Jika semua akun aman)
    journal = models.Journal(
        number=next_journal_number(db, branch_id), 
        date=date_val,
        description=description,
        reference=number_ref,
        source="auto",  
        created_by=user_id,
        branch_id=branch_id 
    )
    db.add(journal)
    db.flush() 

    # 4. Masukkan Baris Jurnal (Entry Lines)
    for entry in entries:
        if entry["debit"] == 0 and entry["credit"] == 0:
            continue 
            
        account = db.query(models.Account).filter(models.Account.code == entry["code"]).first()
            
        if entry["debit"] > 0:
            db.add(models.JournalEntryLine(
                journal_id=journal.id,
                debit_account_id=account.id,
                credit_account_id=None,
                amount=entry["debit"], 
                description=description
            ))
            
        if entry["credit"] > 0:
            db.add(models.JournalEntryLine(
                journal_id=journal.id,
                debit_account_id=None,
                credit_account_id=account.id,
                amount=entry["credit"], 
                description=description
            ))
            
    return journal


def next_cash_number(db: Session, branch_id: int) -> str:
    """Generate nomor transaksi kas unik per cabang"""
    today = get_local_date()
    b_id = branch_id or 0
    prefix = f"KAS-C{b_id}-{today.strftime('%Y%m%d')}"
    last = db.query(models.CashTransaction).filter(
        models.CashTransaction.number.like(f"{prefix}%")
    ).order_by(models.CashTransaction.id.desc()).with_for_update().first()
    seq = int(last.number[-4:]) + 1 if last else 1
    return f"{prefix}{seq:04d}"


def get_account_balance(
    db: Session, 
    account_id: int,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    branch_id: Optional[int] = None
) -> float:
    """Hitung saldo akun berdasarkan jurnal entries, di-filter berdasarkan cabang"""
    account = db.query(models.Account).get(account_id)
    if not account:
        return 0.0

    q_debit = db.query(func.sum(models.JournalEntryLine.amount)).join(models.Journal).filter(
        models.JournalEntryLine.debit_account_id == account_id
    )
    q_credit = db.query(func.sum(models.JournalEntryLine.amount)).join(models.Journal).filter(
        models.JournalEntryLine.credit_account_id == account_id
    )

    if branch_id is not None:
        q_debit = q_debit.filter(models.Journal.branch_id == branch_id)
        q_credit = q_credit.filter(models.Journal.branch_id == branch_id)

    if start_date:
        q_debit = q_debit.filter(models.Journal.date >= start_date)
        q_credit = q_credit.filter(models.Journal.date >= start_date)
    if end_date:
        q_debit = q_debit.filter(models.Journal.date <= end_date)
        q_credit = q_credit.filter(models.Journal.date <= end_date)

    # Saldo awal (Opening Balance) hanya dihitung jika dilihat secara Global (Pusat)
    op_balance = account.opening_balance if not branch_id else 0.0

    total_debit = (q_debit.scalar() or 0) + (op_balance if account.normal_balance == "debit" else 0)
    total_credit = (q_credit.scalar() or 0) + (op_balance if account.normal_balance == "credit" else 0)

    if account.normal_balance == "debit":
        return total_debit - total_credit
    else:
        return total_credit - total_debit


# ══════════════════════════════════════════════════════════════════════════════
# CHART OF ACCOUNTS (Data Master - Tetap Global)
# ══════════════════════════════════════════════════════════════════════════════

class AccountCreate(BaseModel):
    code: str
    name: str
    type: str
    subtype: Optional[str] = None
    normal_balance: str = "debit"
    opening_balance: float = 0.0

class AccountUpdate(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None
    type: Optional[str] = None
    subtype: Optional[str] = None
    normal_balance: Optional[str] = None
    opening_balance: Optional[float] = None
    is_active: Optional[bool] = None


@router.get("/accounts")
def get_accounts(
    type: Optional[str] = None,
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    q = db.query(models.Account)
    if active_only:
        q = q.filter(models.Account.is_active == True)
    if type:
        q = q.filter(models.Account.type == type)
    accounts = q.order_by(models.Account.code).all()
    return [{
        "id": a.id, "code": a.code, "name": a.name,
        "type": a.type, "subtype": a.subtype,
        "normal_balance": a.normal_balance,
        "opening_balance": a.opening_balance,
        "is_active": a.is_active,
        "current_balance": get_account_balance(db, a.id, branch_id=current_user.active_branch_id)
    } for a in accounts]


@router.post("/accounts")
def create_account(
    data: AccountCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if db.query(models.Account).filter(models.Account.code == data.code).first():
        raise HTTPException(400, "Kode akun sudah digunakan")

    valid_types = ["asset", "liability", "equity", "revenue", "expense"]
    if data.type not in valid_types:
        raise HTTPException(400, f"Tipe akun harus salah satu dari: {valid_types}")

    normal = "debit" if data.type in ["asset", "expense"] else "credit"

    # Saldo awal TIDAK diisi lewat editor akun: field opening_balance hanya tampak di laporan
    # global (per-cabang diabaikan di get_account_balance), sehingga nilai non-nol bisa membuat
    # neraca cabang tak balance. Saldo awal yang benar dicatat sebagai JURNAL berimbang lewat
    # menu Setup Saldo Awal / Jurnal Manual.
    if data.opening_balance and abs(float(data.opening_balance)) > 0.005:
        raise HTTPException(400, "Saldo awal tidak diisi di sini. Gunakan menu Setup Saldo Awal atau Jurnal Manual agar tercatat berimbang & per-cabang.")

    obj = models.Account(
        code=data.code, name=data.name, type=data.type,
        subtype=data.subtype, normal_balance=normal,
        opening_balance=0.0
    )
    db.add(obj)
    db.commit()
    db.refresh(obj)
    write_audit(db, current_user.id, "CREATE", "accounts", obj.id, f"Buat akun {obj.code} - {obj.name}")
    db.commit()
    return {"id": obj.id, "code": obj.code, "name": obj.name, "message": "Akun dibuat"}


@router.put("/accounts/{account_id}")
def update_account(
    account_id: int, data: AccountUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    obj = db.query(models.Account).with_for_update().get(account_id)
    if not obj:
        raise HTTPException(404, "Akun tidak ditemukan")
    perubahan = data.model_dump(exclude_unset=True)
    # Lihat catatan di create_account: saldo awal tak boleh di-set dari editor akun.
    if perubahan.get("opening_balance") and abs(float(perubahan["opening_balance"])) > 0.005:
        raise HTTPException(400, "Saldo awal tidak diisi di sini. Gunakan menu Setup Saldo Awal atau Jurnal Manual agar tercatat berimbang & per-cabang.")
    for k, v in perubahan.items():
        setattr(obj, k, v)
    db.commit()
    db.refresh(obj)
    write_audit(db, current_user.id, "UPDATE", "accounts", obj.id, f"Update akun {obj.code}")
    db.commit()
    return {"id": obj.id, "message": "Akun diperbarui"}


@router.delete("/accounts/{account_id}")
def delete_account(
    account_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    obj = db.query(models.Account).with_for_update().get(account_id)
    if not obj:
        raise HTTPException(404, "Akun tidak ditemukan")
    has_transactions = db.query(models.JournalEntryLine).filter(
        (models.JournalEntryLine.debit_account_id == account_id) |
        (models.JournalEntryLine.credit_account_id == account_id)
    ).first()
    if has_transactions:
        obj.is_active = False
        db.commit()
        return {"message": "Akun dinonaktifkan (ada transaksi terkait)"}
    db.delete(obj)
    db.commit()
    return {"message": "Akun dihapus"}


def get_default_accounts() -> list[tuple[str, str, str, str, str]]:
    return [
            # --- ASET ---
            ("1-1100", "Kas", "asset", "current_asset", "debit"),
            ("1-1200", "Bank", "asset", "current_asset", "debit"),
            ("1-1250", "E-Money / Dompet Digital", "asset", "current_asset", "debit"), # BARU: uang masuk via e-money/dompet digital
            ("1-1300", "Piutang Usaha", "asset", "current_asset", "debit"),
            ("1-1400", "Persediaan Barang", "asset", "current_asset", "debit"),
            ("1-1500", "Uang Muka Pembelian", "asset", "current_asset", "debit"),
            ("1-1550", "PPN Masukan (Pajak Dibayar Dimuka)", "asset", "current_asset", "debit"), # BARU
            ("1-1600", "Perlengkapan Toko", "asset", "current_asset", "debit"),
            ("1-2100", "Peralatan Toko", "asset", "fixed_asset", "debit"),
            ("1-2200", "Kendaraan", "asset", "fixed_asset", "debit"),
            ("1-2300", "Bangunan", "asset", "fixed_asset", "debit"),
            
            # --- KEWAJIBAN ---
            ("2-1100", "Hutang Usaha", "liability", "current_liability", "credit"),
            ("2-1200", "Hutang PPN (Keluaran)", "liability", "current_liability", "credit"), # DIUBAH NAMA SEDIKIT
            ("2-1300", "Uang Muka Penjualan", "liability", "current_liability", "credit"),
            ("2-1400", "Beban Masih Harus Dibayar", "liability", "current_liability", "credit"),
            
            # --- EKUITAS ---
            ("3-1100", "Modal Pemilik", "equity", "capital", "credit"),
            ("3-1200", "Prive / Pengambilan Pemilik", "equity", "capital", "debit"),
            ("3-1300", "Laba Ditahan", "equity", "retained_earnings", "credit"),
            ("3-1999", "Modal Transisi (Setup Awal Stok)", "equity", "capital", "credit"),
            ("3-2000", "Mutasi Antar Cabang", "equity", "capital", "credit"),
            ("3-2100", "Transfer dari Pusat (Inter-Branch)", "equity", "capital", "credit"),
            ("3-2200", "Kirim Barang ke Cabang (Inter-Branch)", "equity", "capital", "credit"),
            ("3-2300", "Setoran ke Pusat", "equity", "capital", "credit"),
            ("3-2400", "Setoran dari Cabang", "equity", "capital", "credit"),
            
            # --- PENDAPATAN ---
            ("4-1100", "Penjualan", "revenue", "operating", "credit"),
            ("4-1150", "Diskon Penjualan", "revenue", "operating", "debit"), # BARU
            ("4-1200", "Retur Penjualan", "revenue", "operating", "debit"),
            ("4-1500", "Pendapatan Biaya Lain (Penjualan)", "revenue", "operating", "credit"), # BARU: biaya lain ditagihkan ke pelanggan di kasir
            ("4-1300", "Pendapatan Lain-lain (Surplus Opname)", "revenue", "non_operating", "credit"),
            ("4-1400", "Pendapatan Lain-lain (Penghapusan Saldo)", "revenue", "non_operating", "credit"),
            ("4-2000", "Diskon Pembelian", "revenue", "non_operating", "credit"), # BARU
            
            # --- BEBAN & HPP ---
            ("5-1100", "Harga Pokok Penjualan", "expense", "cogs", "debit"),
            ("5-1200", "Beban Susut & Selisih Persediaan", "expense", "cogs", "debit"), # BARU
            ("5-2100", "Beban Gaji Karyawan", "expense", "operating", "debit"),
            ("5-2200", "Beban Sewa Toko", "expense", "operating", "debit"),
            ("5-2300", "Beban Listrik & Air", "expense", "operating", "debit"),
            ("5-2400", "Beban Perlengkapan", "expense", "operating", "debit"),
            ("5-2500", "Beban Transportasi", "expense", "operating", "debit"),
            ("5-2600", "Beban Pemasaran", "expense", "operating", "debit"),
            ("5-2700", "Beban Lain-lain", "expense", "non_operating", "debit"),
        ]


def ensure_default_accounts(db: Session) -> int:
    created = 0
    for code, name, type_, subtype, normal in get_default_accounts():
        account = db.query(models.Account).filter(models.Account.code == code).first()
        if account:
            if not account.name:
                account.name = name
            if not account.type:
                account.type = type_
            if not account.subtype:
                account.subtype = subtype
            if not account.normal_balance:
                account.normal_balance = normal
            continue

        db.add(models.Account(
            code=code,
            name=name,
            type=type_,
            subtype=subtype,
            normal_balance=normal,
        ))
        created += 1
    db.flush()
    return created


def pastikan_akun_ada(db: Session, codes: list[str]) -> None:
    """Pastikan akun-akun (kode) yang WAJIB dipakai sebuah jurnal otomatis benar-benar ada
    SEBELUM transaksi memutasi stok.

    Tujuan: mencegah jurnal gagal di tengah jalan (mis. akun di-rename/dihapus di COA klien)
    yang akan membuat stok berpindah TANPA jurnal pendamping → GL menyimpang diam-diam.
    Akun yang hilang dibuat ulang dari template COA default (idempotent). Kode non-default
    dibiarkan: bila benar-benar tak ada, create_auto_journal yang akan menolak dengan jelas
    sehingga seluruh transaksi rollback (tidak ada commit sebagian)."""
    template = {row[0]: row for row in get_default_accounts()}  # code -> (code,name,type,subtype,normal)
    perlu_flush = False
    for code in codes:
        if db.query(models.Account.id).filter(models.Account.code == code).first():
            continue
        t = template.get(code)
        if not t:
            continue
        db.add(models.Account(
            code=t[0], name=t[1], type=t[2], subtype=t[3], normal_balance=t[4], is_active=True
        ))
        perlu_flush = True
    if perlu_flush:
        db.flush()


@router.post("/accounts/seed-default")
def seed_default_accounts(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if "admin" not in (current_user.role or ""):
        raise HTTPException(403, "Hanya admin")

    created = ensure_default_accounts(db)
    db.commit()
    if created == 0:
        return {"message": "Chart of accounts sudah lengkap"}
    return {"message": f"{created} akun default dibuat"}


# ══════════════════════════════════════════════════════════════════════════════
# JURNAL UMUM
# ══════════════════════════════════════════════════════════════════════════════

class JournalLineCreate(BaseModel):
    debit_account_id: Optional[int] = None
    credit_account_id: Optional[int] = None
    amount: float
    description: Optional[str] = None

class JournalCreate(BaseModel):
    date: date
    description: str
    reference: Optional[str] = None
    lines: List[JournalLineCreate]


@router.get("/journals")
def get_journals(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    skip: int = 0, limit: int = 100,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    q = get_query(db, models.Journal, current_user)
    if start_date:
        q = q.filter(models.Journal.date >= start_date)
    if end_date:
        q = q.filter(models.Journal.date <= end_date)
    journals = q.order_by(models.Journal.id.desc()).offset(skip).limit(limit).all()

    result = []
    for j in journals:
        total_debit = sum(l.amount for l in j.lines if l.debit_account_id)
        total_credit = sum(l.amount for l in j.lines if l.credit_account_id)
        result.append({
            "id": j.id, "number": j.number,
            "date": str(j.date), "description": j.description,
            "reference": j.reference, "source": j.source,
            "creator": j.creator.username if j.creator else "-",
            "branch_id": j.branch_id,
            "total_debit": total_debit, "total_credit": total_credit,
            "balanced": abs(total_debit - total_credit) < 0.01,
            "lines": [{
                "id": l.id,
                "debit_account": f"{l.debit_account.code} - {l.debit_account.name}" if l.debit_account else None,
                "debit_account_id": l.debit_account_id,
                "credit_account": f"{l.credit_account.code} - {l.credit_account.name}" if l.credit_account else None,
                "credit_account_id": l.credit_account_id,
                "amount": l.amount, "description": l.description
            } for l in j.lines]
        })
    return result


@router.get("/journals/{journal_id}")
def get_journal(
    journal_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    j = get_query(db, models.Journal, current_user).filter(models.Journal.id == journal_id).first()
    if not j:
        raise HTTPException(404, "Jurnal tidak ditemukan")
    return {
        "id": j.id, "number": j.number,
        "date": str(j.date), "description": j.description,
        "reference": j.reference, "source": j.source,
        "creator": j.creator.username if j.creator else "-",
        "branch_id": j.branch_id,
        "lines": [{
            "id": l.id,
            "debit_account_id": l.debit_account_id,
            "debit_account_code": l.debit_account.code if l.debit_account else None,
            "debit_account_name": l.debit_account.name if l.debit_account else None,
            "credit_account_id": l.credit_account_id,
            "credit_account_code": l.credit_account.code if l.credit_account else None,
            "credit_account_name": l.credit_account.name if l.credit_account else None,
            "amount": l.amount, "description": l.description
        } for l in j.lines]
    }


@router.post("/journals")
def create_journal(
    data: JournalCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    total_debit = sum(l.amount for l in data.lines if l.debit_account_id)
    total_credit = sum(l.amount for l in data.lines if l.credit_account_id)

    if abs(total_debit - total_credit) > 0.01:
        raise HTTPException(400,
            f"Jurnal tidak balance. Debit: {total_debit:,.0f}, Credit: {total_credit:,.0f}, "
            f"Selisih: {abs(total_debit - total_credit):,.0f}")

    if not data.lines:
        raise HTTPException(400, "Jurnal harus memiliki minimal 1 baris")

    for line in data.lines:
        if line.debit_account_id:
            if not db.query(models.Account).get(line.debit_account_id):
                raise HTTPException(404, f"Akun debit {line.debit_account_id} tidak ditemukan")
        if line.credit_account_id:
            if not db.query(models.Account).get(line.credit_account_id):
                raise HTTPException(404, f"Akun kredit {line.credit_account_id} tidak ditemukan")
        if line.amount <= 0:
            raise HTTPException(400, "Jumlah harus lebih dari 0")

    b_id = resolve_active_branch_id(db, current_user)
    assert_books_open(db, b_id, data.date, "Jurnal manual")
    number = next_journal_number(db, b_id)
    
    journal = models.Journal(
        number=number, date=data.date,
        description=data.description, reference=data.reference,
        source="manual", created_by=current_user.id,
        branch_id=b_id
    )
    db.add(journal)
    db.flush()

    for line in data.lines:
        db.add(models.JournalEntryLine(
            journal_id=journal.id,
            debit_account_id=line.debit_account_id,
            credit_account_id=line.credit_account_id,
            amount=line.amount,
            description=line.description
        ))

    db.commit()
    db.refresh(journal)
    write_audit(db, current_user.id, "CREATE", "journals", journal.id,
                f"Jurnal {journal.number}: {data.description}")
    db.commit()
    return {"id": journal.id, "number": journal.number, "message": "Jurnal berhasil disimpan"}


@router.delete("/journals/{journal_id}")
def delete_journal(
    journal_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    j = get_query(db, models.Journal, current_user).filter(models.Journal.id == journal_id).first()
    if not j:
        raise HTTPException(404, "Jurnal tidak ditemukan")
    if j.source != "manual":
        raise HTTPException(400, "Jurnal otomatis tidak bisa dihapus")
    assert_books_open(
        db,
        j.branch_id if j.branch_id is not None else current_user.active_branch_id,
        j.date,
        "Hapus jurnal",
    )
    write_audit(db, current_user.id, "DELETE", "journals", j.id, f"Hapus jurnal {j.number}")
    db.delete(j)
    db.commit()
    return {"message": "Jurnal dihapus"}


# ══════════════════════════════════════════════════════════════════════════════
# BUKU BESAR
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/book-close/status")
def get_book_close_status(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    b_id = resolve_active_branch_id(db, current_user)
    locked_until = get_books_locked_until(db, b_id)
    latest = (
        db.query(models.BookClose)
        .filter(models.BookClose.branch_id == b_id, models.BookClose.is_active == True)
        .order_by(models.BookClose.period_end.desc(), models.BookClose.id.desc())
        .first()
    )
    return {
        "branch_id": b_id,
        "locked_until": str(locked_until) if locked_until else None,
        "latest": (
            {
                "id": latest.id,
                "close_type": latest.close_type,
                "period_start": str(latest.period_start),
                "period_end": str(latest.period_end),
                "notes": latest.notes,
                "created_by": latest.created_by,
                "created_at": latest.created_at.isoformat() if latest.created_at else None,
            }
            if latest
            else None
        ),
    }


@router.get("/book-close/history")
def get_book_close_history(
    limit: int = 24,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    b_id = resolve_active_branch_id(db, current_user)
    q = (
        db.query(models.BookClose)
        .filter(models.BookClose.branch_id == b_id)
        .order_by(models.BookClose.period_end.desc(), models.BookClose.id.desc())
        .limit(limit)
        .all()
    )
    return [
        {
            "id": x.id,
            "close_type": x.close_type,
            "period_start": str(x.period_start),
            "period_end": str(x.period_end),
            "notes": x.notes,
            "is_active": bool(x.is_active),
            "created_by": x.created_by,
            "created_at": x.created_at.isoformat() if x.created_at else None,
            "reopened_by": x.reopened_by,
            "reopened_at": x.reopened_at.isoformat() if x.reopened_at else None,
        }
        for x in q
    ]


@router.post("/book-close/close")
def close_books(
    data: BookCloseCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    b_id = resolve_active_branch_id(db, current_user)
    end_date = data.end_date
    if end_date > get_local_date():
        raise HTTPException(400, "Tanggal tutup buku tidak boleh di masa depan.")

    close_type = (data.close_type or "month").strip().lower()
    if close_type not in ("month", "year"):
        raise HTTPException(400, "close_type harus 'month' atau 'year'.")

    if close_type == "month":
        last_day = monthrange(end_date.year, end_date.month)[1]
        if end_date.day != last_day:
            raise HTTPException(
                400,
                "Untuk tutup buku bulanan, tanggal harus di hari terakhir bulan tersebut.",
            )
        period_start = end_date.replace(day=1)
    else:
        if end_date.month != 12 or end_date.day != 31:
            raise HTTPException(400, "Untuk tutup buku tahunan, tanggal harus 31 Desember.")
        period_start = date(end_date.year, 1, 1)

    locked_until = get_books_locked_until(db, b_id)
    if locked_until and end_date <= locked_until:
        raise HTTPException(
            400,
            f"Buku sudah ditutup sampai {locked_until}. Pilih tanggal setelah itu.",
        )

    obj = models.BookClose(
        branch_id=b_id,
        close_type=close_type,
        period_start=period_start,
        period_end=end_date,
        notes=(data.notes or "").strip() or None,
        is_active=True,
        created_by=current_user.id,
    )
    db.add(obj)
    db.commit()
    db.refresh(obj)
    write_audit(
        db,
        current_user.id,
        "CREATE",
        "book_closes",
        obj.id,
        f"Tutup buku ({close_type}) sampai {end_date}",
    )
    db.commit()
    return {"message": f"✓ Buku berhasil ditutup sampai {end_date}", "locked_until": str(end_date)}


@router.post("/book-close/reopen-latest")
def reopen_latest_books(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    b_id = resolve_active_branch_id(db, current_user)
    obj = (
        db.query(models.BookClose)
        .filter(models.BookClose.branch_id == b_id, models.BookClose.is_active == True)
        .order_by(models.BookClose.period_end.desc(), models.BookClose.id.desc())
        .first()
    )
    if not obj:
        raise HTTPException(404, "Belum ada riwayat tutup buku.")

    obj.is_active = False
    obj.reopened_by = current_user.id
    obj.reopened_at = datetime.now(WITA)
    db.commit()

    locked_until = get_books_locked_until(db, b_id)
    write_audit(
        db,
        current_user.id,
        "UPDATE",
        "book_closes",
        obj.id,
        f"Buka tutup buku terakhir (sebelumnya sampai {obj.period_end})",
    )
    db.commit()
    return {
        "message": "Periode tutup buku terakhir dibuka kembali.",
        "locked_until": str(locked_until) if locked_until else None,
    }


@router.get("/ledger/{account_id}")
def get_ledger(
    account_id: int,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    account = db.query(models.Account).get(account_id)
    if not account:
        raise HTTPException(404, "Akun tidak ditemukan")

    q_debit = db.query(models.JournalEntryLine).filter(
        models.JournalEntryLine.debit_account_id == account_id
    ).join(models.Journal)
    q_credit = db.query(models.JournalEntryLine).filter(
        models.JournalEntryLine.credit_account_id == account_id
    ).join(models.Journal)

    b_id = resolve_active_branch_id(db, current_user)
    if b_id:
        q_debit = q_debit.filter(models.Journal.branch_id == b_id)
        q_credit = q_credit.filter(models.Journal.branch_id == b_id)

    if start_date:
        q_debit = q_debit.filter(models.Journal.date >= start_date)
        q_credit = q_credit.filter(models.Journal.date >= start_date)
    if end_date:
        q_debit = q_debit.filter(models.Journal.date <= end_date)
        q_credit = q_credit.filter(models.Journal.date <= end_date)

    entries = []

    for line in q_debit.all():
        entries.append({
            "date": str(line.journal.date),
            "journal_number": line.journal.number,
            "description": line.journal.description,
            "debit": line.amount, "credit": 0,
            "sort_key": str(line.journal.date) + str(line.journal.id).zfill(10)
        })

    for line in q_credit.all():
        entries.append({
            "date": str(line.journal.date),
            "journal_number": line.journal.number,
            "description": line.journal.description,
            "debit": 0, "credit": line.amount,
            "sort_key": str(line.journal.date) + str(line.journal.id).zfill(10)
        })

    entries.sort(key=lambda x: x["sort_key"])

    running_balance = account.opening_balance if not b_id else 0.0
    for e in entries:
        if account.normal_balance == "debit":
            running_balance += e["debit"] - e["credit"]
        else:
            running_balance += e["credit"] - e["debit"]
        e["balance"] = running_balance
        del e["sort_key"]

    total_debit = sum(e["debit"] for e in entries)
    total_credit = sum(e["credit"] for e in entries)
    closing_balance = get_account_balance(db, account_id, start_date, end_date, b_id)

    return {
        "account": {
            "id": account.id, "code": account.code,
            "name": account.name, "type": account.type,
            "normal_balance": account.normal_balance,
            "opening_balance": account.opening_balance
        },
        "entries": entries,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "closing_balance": closing_balance,
        "period": {"start": str(start_date) if start_date else None,
                   "end": str(end_date) if end_date else None}
    }


# ══════════════════════════════════════════════════════════════════════════════
# NERACA SALDO (TRIAL BALANCE)
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/trial-balance")
def get_trial_balance(
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    accounts = db.query(models.Account).filter(
        models.Account.is_active == True
    ).order_by(models.Account.code).all()

    rows = []
    total_debit = 0
    total_credit = 0
    b_id = resolve_active_branch_id(db, current_user)

    for a in accounts:
        balance = get_account_balance(db, a.id, end_date=end_date, branch_id=b_id)

        if balance == 0:
            continue 

        if a.normal_balance == "debit":
            debit_bal = balance if balance >= 0 else 0
            credit_bal = abs(balance) if balance < 0 else 0
        else:
            credit_bal = balance if balance >= 0 else 0
            debit_bal = abs(balance) if balance < 0 else 0

        total_debit += debit_bal
        total_credit += credit_bal

        rows.append({
            "code": a.code, "name": a.name,
            "type": a.type, "subtype": a.subtype,
            "debit": debit_bal, "credit": credit_bal
        })

    return {
        "rows": rows,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "balanced": abs(total_debit - total_credit) < 1,
        "as_of": str(end_date) if end_date else str(get_local_date())
    }


# ══════════════════════════════════════════════════════════════════════════════
# LAPORAN LABA RUGI AKUNTANSI
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/income-statement")
def get_income_statement(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if not start_date:
        start_date = get_local_date().replace(day=1)
    if not end_date:
        end_date = get_local_date()

    accounts = db.query(models.Account).filter(
        models.Account.is_active == True,
        models.Account.type.in_(["revenue", "expense"])
    ).order_by(models.Account.code).all()

    revenues = []           # Pendapatan murni (credit normal)
    contra_revenues = []    # Diskon ke customer, retur (debit normal, type=revenue)
    supplier_discounts = [] # Diskon dari supplier (4-2000)
    expenses_cogs = []
    expenses_operating = []
    expenses_other = []
    b_id = current_user.active_branch_id

    for a in accounts:
        balance = get_account_balance(db, a.id, start_date, end_date, b_id)
        row = {"code": a.code, "name": a.name, "amount": abs(balance)}

        if a.type == "revenue":
            # Akun 4-2000 Diskon Pembelian = pendapatan dari supplier, pisahkan
            if a.code == "4-2000":
                supplier_discounts.append(row)
            # Contra revenue: normal_balance debit (Diskon Penjualan, Retur)
            elif a.normal_balance == "debit":
                contra_revenues.append(row)
            else:
                revenues.append(row)
        elif a.type == "expense":
            if a.subtype == "cogs":
                expenses_cogs.append(row)
            elif a.subtype == "operating":
                expenses_operating.append(row)
            else:
                expenses_other.append(row)

    # ── Kalkulasi bertahap ──────────────────────────────────────────────────
    gross_revenue          = sum(r["amount"] for r in revenues)
    total_contra_revenue   = sum(r["amount"] for r in contra_revenues)   # diskon ke customer
    total_supplier_disc    = sum(r["amount"] for r in supplier_discounts) # diskon dari supplier

    # Net Revenue = Penjualan - Diskon ke Customer - Retur
    net_revenue = gross_revenue - total_contra_revenue

    # Gross Profit = Net Revenue - (HPP - Diskon Supplier)
    total_cogs  = sum(e["amount"] for e in expenses_cogs)
    net_cogs    = total_cogs - total_supplier_disc  # diskon supplier kurangi HPP
    gross_profit = net_revenue - net_cogs

    total_operating_expense = sum(e["amount"] for e in expenses_operating)
    operating_profit        = gross_profit - total_operating_expense
    total_other_expense     = sum(e["amount"] for e in expenses_other)
    net_profit              = operating_profit - total_other_expense

    return {
        "period": {"start": str(start_date), "end": str(end_date)},

        # Detail Pendapatan
        "revenues": revenues,
        "gross_revenue": gross_revenue,
        "contra_revenues": contra_revenues,          # 👈 BARU (Diskon Penjualan, Retur)
        "total_contra_revenue": total_contra_revenue,# 👈 BARU
        "total_revenue": net_revenue,                # Net setelah diskon

        # Detail HPP
        "cogs": expenses_cogs,
        "supplier_discounts": supplier_discounts,    # 👈 BARU (Diskon dari Supplier)
        "total_supplier_discount": total_supplier_disc, # 👈 BARU
        "total_cogs": net_cogs,                      # Net setelah diskon supplier
        "gross_cogs": total_cogs,

        "gross_profit": gross_profit,
        "gross_margin_percent": round(gross_profit / net_revenue * 100, 2) if net_revenue else 0,

        # Beban
        "operating_expenses": expenses_operating,
        "total_operating_expense": total_operating_expense,
        "operating_profit": operating_profit,
        "other_expenses": expenses_other,
        "total_other_expense": total_other_expense,
        "net_profit": net_profit
    }


# ══════════════════════════════════════════════════════════════════════════════
# NERACA (BALANCE SHEET)
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/balance-sheet")
def get_balance_sheet(
    as_of: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if not as_of:
        as_of = get_local_date()

    accounts = db.query(models.Account).filter(
        models.Account.is_active == True,
        models.Account.type.in_(["asset", "liability", "equity"])
    ).order_by(models.Account.code).all()

    assets_current = []
    assets_fixed = []
    liabilities_current = []
    liabilities_long = []
    equity_items = []
    b_id = current_user.active_branch_id

    for a in accounts:
        balance = get_account_balance(db, a.id, end_date=as_of, branch_id=b_id)
        
        # 🛡️ FIX: Sesuaikan tanda (sign) agar Liability & Equity dijumlahkan berdasarkan saldo KREDIT
        # Jika akun bertipe Liability/Equity tapi saldo normalnya DEBIT (seperti Prive atau Setoran), 
        # maka kita balik tandanya agar dia mengurangi total (contra-account).
        display_balance = balance
        if a.type in ["liability", "equity"] and a.normal_balance == "debit":
            display_balance = -balance
            
        row = {"code": a.code, "name": a.name, "amount": display_balance}

        if a.type == "asset":
            if a.subtype == "fixed_asset":
                assets_fixed.append(row)
            else:
                assets_current.append(row)
        elif a.type == "liability":
            if a.subtype == "long_term_liability":
                liabilities_long.append(row)
            else:
                liabilities_current.append(row)
        elif a.type == "equity":
            equity_items.append(row)

    year_start = as_of.replace(month=1, day=1)
    revenue_accounts = db.query(models.Account).filter(
        models.Account.is_active == True,
        models.Account.type == "revenue"
    ).all()
    expense_accounts = db.query(models.Account).filter(
        models.Account.is_active == True,
        models.Account.type == "expense"
    ).all()

    ytd_revenue = sum(get_account_balance(db, a.id, year_start, as_of, b_id) for a in revenue_accounts)
    ytd_expense = sum(get_account_balance(db, a.id, year_start, as_of, b_id) for a in expense_accounts)
    current_period_profit = ytd_revenue - ytd_expense

    # ── Laba Ditahan dinamis (akumulasi laba/rugi tahun-tahun SEBELUM tahun berjalan) ──
    # Aset/kewajiban/ekuitas dijumlahkan kumulatif sejak awal, tetapi laba tahun-tahun lalu
    # tidak pernah dipindah ke akun ekuitas karena sistem tidak memposting jurnal penutup
    # (Tutup Buku hanya mengunci periode). Tanpa baris ini, neraca akan meleset sebesar laba
    # akumulasi sebelum 1 Januari tahun berjalan begitu data sudah lintas tahun. Dihitung live
    # dari baris jurnal → konsisten dengan restatement Fase 3 (koreksi back-date ikut terhitung)
    # dan tetap balance walau buku dibuka-tutup berkali-kali.
    batas_tahun_lalu = year_start - timedelta(days=1)
    laba_ditahan_akumulasi = (
        sum(get_account_balance(db, a.id, None, batas_tahun_lalu, b_id) for a in revenue_accounts)
        - sum(get_account_balance(db, a.id, None, batas_tahun_lalu, b_id) for a in expense_accounts)
    )
    if abs(laba_ditahan_akumulasi) >= 1:
        equity_items.append({
            "code": "3-1300*",
            "name": f"Laba Ditahan (s/d {year_start.year - 1})",
            "amount": laba_ditahan_akumulasi
        })

    if current_period_profit != 0:
        equity_items.append({
            "code": "3-9999",
            "name": f"Laba/Rugi Periode Berjalan ({as_of.year})",
            "amount": current_period_profit
        })

    total_current_assets = sum(r["amount"] for r in assets_current)
    total_fixed_assets = sum(r["amount"] for r in assets_fixed)
    total_assets = total_current_assets + total_fixed_assets

    total_current_liabilities = sum(r["amount"] for r in liabilities_current)
    total_long_liabilities = sum(r["amount"] for r in liabilities_long)
    total_liabilities = total_current_liabilities + total_long_liabilities

    total_equity = sum(r["amount"] for r in equity_items)
    total_liabilities_equity = total_liabilities + total_equity

    return {
        "as_of": str(as_of),
        "assets": {
            "current": assets_current,
            "total_current": total_current_assets,
            "fixed": assets_fixed,
            "total_fixed": total_fixed_assets,
            "total": total_assets
        },
        "liabilities": {
            "current": liabilities_current,
            "total_current": total_current_liabilities,
            "long_term": liabilities_long,
            "total_long_term": total_long_liabilities,
            "total": total_liabilities
        },
        "equity": {
            "items": equity_items,
            "total": total_equity
        },
        "total_liabilities_equity": total_liabilities_equity,
        "balanced": abs(total_assets - total_liabilities_equity) < 1,
        "current_period_profit": current_period_profit
    }


@router.post("/reconcile-inventory-value")
def reconcile_inventory_value(
    dry_run: bool = True,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    """Samakan saldo GL Persediaan (1-1400) dengan Σ nilai lapisan batch FIFO, PER CABANG.

    Selisih (artefak setup/migrasi) dirapikan ke Modal Transisi (3-1999) → TIDAK menyentuh
    laba-rugi. `dry_run=True` (default) hanya menampilkan angka; tanpa dry_run → posting jurnal.
    Idempoten: cabang dengan |selisih| < 1 dilewati, jadi aman dijalankan ulang (setelah
    Bagian A aktif, pembelian tak menambah selisih sehingga ini tetap ~0)."""
    akun = db.query(models.Account).filter(models.Account.code == "1-1400").first()
    if not akun:
        raise HTTPException(400, "Akun Persediaan (1-1400) tidak ditemukan.")
    if not dry_run:
        pastikan_akun_ada(db, ["1-1400", "3-1999"])

    hasil = []
    for br in db.query(models.Branch).order_by(models.Branch.id).all():
        gl = float(get_account_balance(db, akun.id, branch_id=br.id))
        wh_ids = [w.id for w in db.query(models.Warehouse.id).filter(
            models.Warehouse.branch_id == br.id).all()]
        nilai_batch = 0.0
        if wh_ids:
            nilai_batch = float(
                db.query(func.coalesce(func.sum(
                    models.StockBatch.qty_remaining * models.StockBatch.unit_cost), 0.0))
                .filter(models.StockBatch.warehouse_id.in_(wh_ids)).scalar() or 0.0
            )
        selisih = round(gl - nilai_batch, 2)
        baris = {
            "branch_id": br.id, "branch": br.name,
            "gl_persediaan": round(gl, 2), "nilai_batch": round(nilai_batch, 2),
            "selisih": selisih, "diposting": False, "journal_number": None,
        }
        if not dry_run and abs(selisih) >= 1:
            # selisih>0 → GL lebih tinggi: turunkan Persediaan, lawan ke Modal Transisi.
            if selisih > 0:
                entries = [
                    {"code": "1-1400", "debit": 0, "credit": selisih},
                    {"code": "3-1999", "debit": selisih, "credit": 0},
                ]
            else:
                entries = [
                    {"code": "1-1400", "debit": -selisih, "credit": 0},
                    {"code": "3-1999", "debit": 0, "credit": -selisih},
                ]
            j = create_auto_journal(
                db=db, date_val=get_local_date(),
                number_ref=f"RECON-INV-C{br.id}",
                description=f"Penyetaraan nilai persediaan (GL vs batch FIFO) - {br.name}",
                entries=entries, user_id=current_user.id, branch_id=br.id,
                allow_closed_period=True,
            )
            baris["diposting"] = True
            baris["journal_number"] = j.number
            write_audit(db, current_user.id, "CREATE", "journals", j.id,
                        f"Penyetaraan persediaan C{br.id}: {selisih}")
        hasil.append(baris)

    if not dry_run:
        db.commit()
    return {"dry_run": dry_run, "hasil": hasil}


# ══════════════════════════════════════════════════════════════════════════════
# 👇 FITUR BARU: LAPORAN ARUS KAS (CASH FLOW STATEMENT) 👇
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/cash-flow")
def get_cash_flow_statement(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if not start_date:
        start_date = get_local_date().replace(day=1)
    if not end_date:
        end_date = get_local_date()
        
    b_id = current_user.active_branch_id

    # 1. Identifikasi Akun Kas & Bank
    cash_accounts = db.query(models.Account).filter(
        models.Account.code.in_(["1-1100", "1-1200"])
    ).all()
    cash_acc_ids = [a.id for a in cash_accounts]

    # 2. Hitung Saldo Awal Kas (Sebelum periode yang dipilih)
    opening_balance = 0
    prev_day = start_date - timedelta(days=1)
    for acc in cash_accounts:
        opening_balance += get_account_balance(db, acc.id, end_date=prev_day, branch_id=b_id)

    # 3. Tarik Semua Jurnal di Periode Terpilih
    journals = get_query(db, models.Journal, current_user).filter(
        models.Journal.date >= start_date,
        models.Journal.date <= end_date
    ).all()

    acc_impacts = {} 

    # 4. Bedah Jurnal untuk mencari lawan dari akun kas
    for j in journals:
        touches_cash = False
        net_cash_change = 0
        
        for line in j.lines:
            if line.debit_account_id in cash_acc_ids:
                touches_cash = True
                net_cash_change += line.amount
            if line.credit_account_id in cash_acc_ids:
                touches_cash = True
                net_cash_change -= line.amount
                
        if touches_cash and net_cash_change != 0:
            # Pasangan HPP↔Persediaan pada jurnal PENJUALAN bersifat non-kas & saling meniadakan
            # (Dr 5-1100 = Cr 1-1400) → diabaikan agar arus kas penjualan = nilai uang masuk saja.
            # TAPI pembelian persediaan TUNAI (Dr 1-1400 / Cr Kas) TIDAK punya 5-1100 — di situ
            # 1-1400 adalah lawan kas yang sah, jadi JANGAN diabaikan; kalau diabaikan, arus keluar
            # pembelian hilang & saldo kas akhir tak rekonsiliasi.
            jurnal_punya_hpp = any(
                (l.debit_account and l.debit_account.code == "5-1100")
                or (l.credit_account and l.credit_account.code == "5-1100")
                for l in j.lines
            )
            for line in j.lines:
                debit_code = line.debit_account.code if line.debit_account else None
                credit_code = line.credit_account.code if line.credit_account else None

                # Abaikan Persediaan/HPP HANYA pada jurnal ber-HPP (pola penjualan).
                if jurnal_punya_hpp and (
                    debit_code in ["1-1400", "5-1100"] or credit_code in ["1-1400", "5-1100"]
                ):
                    continue

                if line.debit_account_id and line.debit_account_id not in cash_acc_ids:
                    acc_impacts[line.debit_account_id] = acc_impacts.get(line.debit_account_id, 0) - line.amount
                if line.credit_account_id and line.credit_account_id not in cash_acc_ids:
                    acc_impacts[line.credit_account_id] = acc_impacts.get(line.credit_account_id, 0) + line.amount

    # 5. Kategorikan ke Aktivitas Standar Akuntansi
    operating_inflows = []
    operating_outflows = []
    investing = []
    financing = []

    for acc_id, impact in acc_impacts.items():
        if abs(impact) < 0.01: continue
        
        acc = db.query(models.Account).get(acc_id)
        if not acc: continue
        row = {"code": acc.code, "name": acc.name, "amount": impact}
        
        if acc.type in ["revenue", "expense"] or (acc.type in ["asset", "liability"] and acc.subtype in ["current_asset", "current_liability"]):
            if impact > 0: operating_inflows.append(row)
            else: operating_outflows.append(row)
        elif acc.type == "asset" and acc.subtype == "fixed_asset":
            investing.append(row)
        elif acc.type == "equity" or (acc.type == "liability" and acc.subtype == "long_term_liability"):
            financing.append(row)

    net_operating = sum(r["amount"] for r in operating_inflows) + sum(r["amount"] for r in operating_outflows)
    net_investing = sum(r["amount"] for r in investing)
    net_financing = sum(r["amount"] for r in financing)
    
    net_cash_flow = net_operating + net_investing + net_financing
    closing_balance = opening_balance + net_cash_flow

    return {
        "period": {"start": str(start_date), "end": str(end_date)},
        "opening_balance": opening_balance,
        "operating": {"inflows": operating_inflows, "outflows": operating_outflows, "net": net_operating},
        "investing": {"items": investing, "net": net_investing},
        "financing": {"items": financing, "net": net_financing},
        "net_cash_flow": net_cash_flow,
        "closing_balance": closing_balance
    }


# ══════════════════════════════════════════════════════════════════════════════
# KAS MASUK / KAS KELUAR 
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/cash-transactions", response_model=list[schemas.CashTransactionOut])
def get_cash_transactions(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    type: Optional[str] = None,
    skip: int = 0, limit: int = 100,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    q = get_query(db, models.CashTransaction, current_user)
    if start_date:
        q = q.filter(models.CashTransaction.date >= start_date)
    if end_date:
        q = q.filter(models.CashTransaction.date <= end_date)
    if type:
        q = q.filter(models.CashTransaction.type == type)
    return q.order_by(models.CashTransaction.id.desc()).offset(skip).limit(limit).all()

@router.post("/cash-transactions", response_model=schemas.CashTransactionOut)
def create_cash_transaction(
    data: schemas.CashTransactionCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if not data.account_id:
        raise HTTPException(400, "Akun (CoA) sumber/tujuan wajib dipilih agar jurnal terbentuk!")

    b_id = current_user.active_branch_id
    assert_books_open(db, b_id, data.date, "Transaksi kas")
    if not data.number:
        data.number = next_cash_number(db, b_id)
        
    obj = models.CashTransaction(
        **data.model_dump(), 
        branch_id=b_id
    )
    db.add(obj)
    db.flush() 

    akun_pasangan = db.query(models.Account).get(data.account_id)
    if not akun_pasangan:
        raise HTTPException(404, "Akun yang dipilih tidak ditemukan")

    # 👇 PENJAGA GERBANG: Cegah Jurnal "Hantu" (Kas vs Kas / Kas vs Bank) 👇
    if akun_pasangan.code in ["1-1100", "1-1200"]:
        raise HTTPException(
            status_code=400, 
            detail="DITOLAK: Anda tidak boleh memilih akun Kas/Bank sebagai lawan transaksi Kas! Pilih akun Beban, Pendapatan, Modal, atau Hutang."
        )
    # 👆 BATAS PENJAGA GERBANG 👆

    jurnal_entries = []
    if data.type == "income": 
        jurnal_entries.append({"code": "1-1100", "debit": data.amount, "credit": 0})
        jurnal_entries.append({"code": akun_pasangan.code, "debit": 0, "credit": data.amount})
    else: 
        jurnal_entries.append({"code": akun_pasangan.code, "debit": data.amount, "credit": 0})
        jurnal_entries.append({"code": "1-1100", "debit": 0, "credit": data.amount})

    create_auto_journal(
        db=db,
        date_val=data.date,
        number_ref=obj.number,
        description=f"Transaksi Kas: {data.description}",
        entries=jurnal_entries,
        user_id=current_user.id,
        branch_id=b_id
    )

    db.commit()
    db.refresh(obj)
    return obj
@router.get("/cash-balance")
def get_cash_balance(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    income = get_query(db, models.CashTransaction, current_user).filter(
        models.CashTransaction.type == "in"
    ).with_entities(func.sum(models.CashTransaction.amount)).scalar() or 0
    
    expense = get_query(db, models.CashTransaction, current_user).filter(
        models.CashTransaction.type == "out"
    ).with_entities(func.sum(models.CashTransaction.amount)).scalar() or 0
    
    sales_cash = get_query(db, models.Sale, current_user).filter(
        models.Sale.status.in_(["paid", "partial"]),
        models.Sale.payment_method == "cash"
    ).with_entities(func.sum(models.Sale.paid)).scalar() or 0
    
    purchase_paid = get_query(db, models.Purchase, current_user).filter(
        models.Purchase.status.in_(["paid", "partial"])
    ).with_entities(func.sum(models.Purchase.paid)).scalar() or 0
    
    return {
        "total_income": income + sales_cash,
        "total_expense": expense + purchase_paid,
        "balance": (income + sales_cash) - (expense + purchase_paid)
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL AKUNTANSI
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/liquid-balances")
def get_liquid_balances(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    b_id = resolve_active_branch_id(db, current_user)
    cash_acc = db.query(models.Account).filter(models.Account.code == "1-1100").first()
    bank_acc = db.query(models.Account).filter(models.Account.code == "1-1200").first()

    cash_balance = get_account_balance(db, cash_acc.id, branch_id=b_id) if cash_acc else 0.0
    bank_balance = get_account_balance(db, bank_acc.id, branch_id=b_id) if bank_acc else 0.0

    return {
        "branch_id": b_id,
        "cash_balance": float(cash_balance or 0.0),
        "bank_balance": float(bank_balance or 0.0),
        "total_available": float((cash_balance or 0.0) + (bank_balance or 0.0)),
    }


@router.get("/export/journal")
def export_journal_excel(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io

    if not start_date:
        start_date = get_local_date().replace(day=1)
    if not end_date:
        end_date = get_local_date()

    journals = get_query(db, models.Journal, current_user).filter(
        models.Journal.date >= start_date,
        models.Journal.date <= end_date
    ).order_by(models.Journal.date, models.Journal.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jurnal Umum"

    hfill = PatternFill("solid", fgColor="1E293B")
    hfont = Font(color="10B981", bold=True, size=10)
    thin = Border(
        left=Side(style="thin", color="334155"),
        right=Side(style="thin", color="334155"),
        top=Side(style="thin", color="334155"),
        bottom=Side(style="thin", color="334155")
    )

    ws.merge_cells("A1:H1")
    ws["A1"] = f"JURNAL UMUM — {start_date.strftime('%d/%m/%Y')} s/d {end_date.strftime('%d/%m/%Y')}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["No. Jurnal", "Tanggal", "Keterangan", "Referensi", "Cabang ID",
               "Akun Debit", "Akun Kredit", "Jumlah (Rp)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    row = 4
    for j in journals:
        for line in j.lines:
            ws.cell(row=row, column=1, value=j.number).border = thin
            ws.cell(row=row, column=2, value=str(j.date)).border = thin
            ws.cell(row=row, column=3, value=j.description).border = thin
            ws.cell(row=row, column=4, value=j.reference or "").border = thin
            ws.cell(row=row, column=5, value=j.branch_id or "Pusat").border = thin
            debit_name = f"{line.debit_account.code} - {line.debit_account.name}" if line.debit_account else ""
            credit_name = f"{line.credit_account.code} - {line.credit_account.name}" if line.credit_account else ""
            ws.cell(row=row, column=6, value=debit_name).border = thin
            ws.cell(row=row, column=7, value=credit_name).border = thin
            amt_cell = ws.cell(row=row, column=8, value=line.amount)
            amt_cell.number_format = '#,##0'
            amt_cell.border = thin
            row += 1

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    cbg = current_user.active_branch_id or "Semua_Cabang"
    fname = f"jurnal_Cabang{cbg}_{start_date}_{end_date}.xlsx"
    
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )


# ══════════════════════════════════════════════════════════════════════════════
# FITUR BARU: ONBOARDING SETUP SALDO AWAL (GATEKEEPER)
# ══════════════════════════════════════════════════════════════════════════════

class SetupBalanceIn(BaseModel):
    cash: float = 0.0
    bank: float = 0.0
    inventory: float = 0.0
    equipment: float = 0.0
    building: float = 0.0
    payable: float = 0.0

    # Optional legacy-compatible detail rows from onboarding UI.
    # Frontend may send extra fields; we only pick the ones we need.
    payable_detail: Optional[List[dict]] = None


class _SetupPayableRow(BaseModel):
    supplier_name: Optional[str] = None
    description: Optional[str] = None
    amount: float = 0.0


@router.get("/opening-inventory-value")
def get_opening_inventory_value(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    branch_id = resolve_active_branch_id(db, current_user)
    return get_branch_inventory_snapshot(db, branch_id)


@router.get("/setup-status")
def get_setup_status(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    """Mengecek apakah cabang aktif saat ini sudah setup saldo awal"""
    b_id = current_user.active_branch_id

    # 🛡️ FIX: Jika belum ada cabang aktif (Admin baru), cek cabang pertama (Pusat)
    if not b_id:
        first_branch = db.query(models.Branch).order_by(models.Branch.id).first()
        if not first_branch:
            return {"is_setup_completed": False} # Belum ada cabang sama sekali

        # Cek status asli cabang pertama
        status = getattr(first_branch, 'is_setup_complete', False)
        return {"is_setup_completed": status}

    branch = db.query(models.Branch).get(b_id)
    if not branch:
        first_branch = db.query(models.Branch).order_by(models.Branch.id).first()
        if not first_branch:
            raise HTTPException(400, "Sistem belum memiliki cabang sama sekali.")
        b_id = first_branch.id
        branch = db.query(models.Branch).get(b_id)
        if not branch:
            raise HTTPException(400, "Cabang aktif tidak ditemukan. Silakan restart aplikasi.")

    # 🏢 CABANG NON-PUSAT GAUSAH ONBOARDING
    # Cabang gausah setup saldo awal jika bukan ID 1 dan statusnya bukan 'Toko Utama'
    if branch.id != 1 and branch.status != "Toko Utama":
        return {"is_setup_completed": True}

    if not branch:
        return {"is_setup_completed": False}

    # 🛡️ FIX TYPO: Sesuai dengan nama kolom di models.py (tanpa 'd')
    status = getattr(branch, 'is_setup_complete', False)
    return {"is_setup_completed": status}


class PkpStatusIn(BaseModel):
    is_pkp: bool
    tarif_ppn: Optional[float] = None


def _toko_pusat(db: Session):
    """Toko Pusat = penyimpan status company-wide (PKP). Fallback ke cabang pertama."""
    return db.query(models.Branch).get(1) or db.query(models.Branch).order_by(models.Branch.id).first()


@router.get("/pkp-status")
def get_pkp_status(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    """Status PKP toko (company-wide, disimpan di Toko Pusat). Saat ON, PPN pembelian dipisah &
    kasir memungut PPN (harga jual dianggap SUDAH termasuk PPN tarif `tarif_ppn`)."""
    branch = _toko_pusat(db)
    return {
        "is_pkp": bool(getattr(branch, "is_pkp", False)) if branch else False,
        "tarif_ppn": float(getattr(branch, "tarif_ppn", 11) or 11) if branch else 11.0,
    }


@router.post("/pkp-status")
def set_pkp_status(
    data: PkpStatusIn,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Nyalakan/matikan mode PKP. Saat ON, PPN pada pembelian BARU dipisah ke PPN Masukan
    (1-1550) dan kasir memungut PPN penjualan (harga termasuk PPN). Pembelian/penjualan lama
    tidak berubah. Hanya admin."""
    if "admin" not in (current_user.role or ""):
        raise HTTPException(403, "Hanya admin yang boleh mengubah status PKP.")
    branch = _toko_pusat(db)
    if not branch:
        raise HTTPException(400, "Sistem belum memiliki cabang.")
    branch.is_pkp = bool(data.is_pkp)
    if data.tarif_ppn is not None:
        branch.tarif_ppn = float(data.tarif_ppn)
    db.add(branch)
    db.commit()
    return {"is_pkp": bool(branch.is_pkp), "tarif_ppn": float(branch.tarif_ppn or 0)}


@router.get("/ppn-report")
def get_ppn_report(
    start: date,
    end: date,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Ringkasan PPN satu periode untuk cabang aktif (untuk lapor SPT Masa PPN):
    - PPN Keluaran = PPN yang dipungut saat JUAL → gerakan bersih akun 2-1200.
    - PPN Masukan  = PPN yang dibayar saat BELI → gerakan bersih akun 1-1550.
    - Kurang Bayar (disetor) = Keluaran − Masukan bila positif; selisih negatif = Lebih Bayar."""
    b_id = resolve_active_branch_id(db, current_user)
    masukan_acc = db.query(models.Account).filter(models.Account.code == "1-1550").first()
    keluaran_acc = db.query(models.Account).filter(models.Account.code == "2-1200").first()
    ppn_masukan = float(get_account_balance(db, masukan_acc.id, start_date=start, end_date=end, branch_id=b_id)) if masukan_acc else 0.0
    ppn_keluaran = float(get_account_balance(db, keluaran_acc.id, start_date=start, end_date=end, branch_id=b_id)) if keluaran_acc else 0.0
    net = ppn_keluaran - ppn_masukan
    return {
        "start": str(start),
        "end": str(end),
        "branch_id": b_id,
        "ppn_keluaran": round(ppn_keluaran, 2),
        "ppn_masukan": round(ppn_masukan, 2),
        "kurang_bayar": round(net, 2) if net > 0 else 0.0,
        "lebih_bayar": round(-net, 2) if net < 0 else 0.0,
    }


@router.post("/setup-initial-balance")
def setup_initial_balance(
    data: SetupBalanceIn,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    b_id = resolve_active_branch_id(db, current_user)
    import uuid
    
    # 🛡️ FIX: Jika user belum milih cabang, paksa tembak ke cabang pertama (Pusat)
    if not b_id:
        first_branch = db.query(models.Branch).order_by(models.Branch.id).first()
        if not first_branch:
            raise HTTPException(400, "Sistem belum memiliki cabang sama sekali.")
        b_id = first_branch.id

    branch = db.query(models.Branch).get(b_id)
    if getattr(branch, 'is_setup_complete', False):
        raise HTTPException(400, "Setup saldo awal sudah pernah dilakukan. Gunakan menu Jurnal Manual jika ingin mengubah.")

    # 🛡️ AUTO-GENERATE CoA: Jika akun belum digenerate, bantu generate otomatis!
    ensure_default_accounts(db)

    required_account_codes = ["1-1100", "1-1200", "1-1400", "1-2100", "1-2300", "2-1100", "3-1100"]
    missing_codes = [
        code
        for code in required_account_codes
        if not db.query(models.Account.id).filter(models.Account.code == code).first()
    ]
    if missing_codes:
        raise HTTPException(400, f"Master Akun (CoA) belum lengkap: {', '.join(missing_codes)}.")

    def next_opening_purchase_number() -> str:
        today_str = datetime.now(WITA).strftime("%Y%m%d")
        prefix = f"OPAP-C{b_id}-{today_str}"
        last = (
            db.query(models.Purchase)
            .filter(models.Purchase.number.like(f"{prefix}%"))
            .order_by(models.Purchase.id.desc())
            .with_for_update()
            .first()
        )
        seq = int(last.number[-4:]) + 1 if last else 1
        return f"{prefix}{seq:04d}"

    def ensure_supplier_id(name: str) -> int:
        cleaned = (name or "").strip()
        if not cleaned:
            raise HTTPException(400, "Nama supplier untuk hutang tidak boleh kosong.")

        existing = (
            db.query(models.Supplier)
            .filter(func.lower(models.Supplier.name) == cleaned.lower())
            .first()
        )
        if existing:
            return existing.id

        sup_code = f"SUP-{uuid.uuid4().hex[:5].upper()}"
        obj = models.Supplier(code=sup_code, name=cleaned, is_active=True)
        db.add(obj)
        db.flush()
        return obj.id

    inventory_snapshot = get_branch_inventory_snapshot(db, b_id)
    inventory_amount = float(data.inventory or 0)
    if inventory_amount <= 0 and inventory_snapshot["inventory_value"] > 0:
        inventory_amount = inventory_snapshot["inventory_value"]

    entries = []
    total_debit = 0.0
    total_credit = 0.0

    # 1. Harta (Debit)
    if data.cash > 0:
        entries.append({"code": "1-1100", "debit": data.cash, "credit": 0})
        total_debit += data.cash
    if data.bank > 0:
        entries.append({"code": "1-1200", "debit": data.bank, "credit": 0})
        total_debit += data.bank
    if inventory_amount > 0:
        entries.append({"code": "1-1400", "debit": inventory_amount, "credit": 0})
        total_debit += inventory_amount
    if data.equipment > 0:
        entries.append({"code": "1-2100", "debit": data.equipment, "credit": 0})
        total_debit += data.equipment
    if data.building > 0:
        entries.append({"code": "1-2300", "debit": data.building, "credit": 0})
        total_debit += data.building

    # 2. Hutang onboarding tidak dimasukkan ke jurnal ini.
    # Hutang akan dibuat sebagai Purchase unpaid agar bisa dibayar di menu Pembelian.

    # 3. Hitung Ekuitas/Modal Pemilik (Otomatis seimbang)
    equity = total_debit - total_credit
    if equity > 0:
        entries.append({"code": "3-1100", "debit": 0, "credit": equity})
    elif equity < 0:
        entries.append({"code": "3-1100", "debit": abs(equity), "credit": 0})

    # 4. Tembakkan ke mesin Jurnal Otomatis FPOS
    if total_debit > 0 or total_credit > 0:
        number = next_journal_number(db, b_id)
        create_auto_journal(
            db=db, 
            date_val=get_local_date(), 
            number_ref=number,
            description="[SISTEM] Setup Saldo Awal Toko",
            entries=entries, 
            user_id=current_user.id, 
            branch_id=b_id
        )

    # 5. Buat hutang onboarding sebagai "Purchase unpaid" + jurnal pembuka,
    # supaya bisa dibayar di menu Pembelian (Bayar Hutang Supplier).
    payable_rows: list[_SetupPayableRow] = []
    if data.payable_detail:
        for raw in data.payable_detail:
            try:
                row = _SetupPayableRow.model_validate(raw)
            except Exception:
                continue
            if row.amount and row.amount > 0:
                payable_rows.append(row)
    elif data.payable and data.payable > 0:
        payable_rows.append(
            _SetupPayableRow(
                supplier_name="Supplier Hutang Awal",
                description="Hutang awal (tanpa rincian)",
                amount=data.payable,
            )
        )

    for row in payable_rows:
        supplier_name = (row.supplier_name or "").strip()
        description = (row.description or "").strip()
        amount = float(row.amount or 0)
        if amount <= 0:
            continue

        # Back-compat: frontend lama kirim {description, amount} saja.
        if not supplier_name:
            supplier_name = description or "Supplier Hutang Awal"

        supplier_id = ensure_supplier_id(supplier_name)
        pur_number = next_opening_purchase_number()

        purchase = models.Purchase(
            number=pur_number,
            date=get_local_date(),
            branch_id=b_id,
            supplier_id=supplier_id,
            subtotal=amount,
            discount=0,
            tax=0,
            total=amount,
            paid=0,
            status="unpaid",
            notes=f"[ONBOARDING] {description}".strip() if description else "[ONBOARDING] Hutang awal",
            created_by=current_user.id,
            created_at=datetime.now(WITA),
        )
        db.add(purchase)
        db.flush()

        create_auto_journal(
            db=db,
            date_val=get_local_date(),
            number_ref=pur_number,
            description=f"[SISTEM] Hutang Awal Supplier: {supplier_name}",
            entries=[
                {"code": "3-1100", "debit": amount, "credit": 0},
                {"code": "2-1100", "debit": 0, "credit": amount},
            ],
            user_id=current_user.id,
            branch_id=b_id,
        )

    # 7. Pastikan pelanggan 'Umum' tersedia
    umum_cust = db.query(models.Customer).filter(func.lower(models.Customer.name) == "umum").first()
    if not umum_cust:
        new_umum = models.Customer(
            code="CUST-UMUM",
            name="Umum",
            phone="-",
            is_active=True
        )
        db.add(new_umum)
        db.flush()

    # 6. Kunci Cabang agar tidak bisa di-setup dua kali
    # 🔥 FIX: Pastikan nama kolom sesuai dengan models.py (is_setup_complete)
    branch.is_setup_complete = True
    
    # 🛡️ FIX: Ikat akun Admin ini ke cabang yang baru di-setup agar tidak bug
    current_user.active_branch_id = b_id
    
    db.commit()

    return {"message": "Setup berhasil! Neraca awal Anda sudah seimbang."}
