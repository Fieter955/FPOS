"""Bangun aset seed FPOS yang deterministik dari empat workbook ekspor iPos.

Workbook sumber sengaja hanya dipakai pada saat pengembangan/build. Runtime FPOS
memuat JSON gzip yang lebih kecil dan lebih cepat, sekaligus sudah tervalidasi.
"""
from __future__ import annotations

import argparse
import gzip
import json
import math
import warnings
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SEED_VERSION = "ipos-master-2026-08-20-v1"
SOURCE_FILES = {
    "S": "satuan harga.xlsx",
    "L": "berdasarkan level harga.xlsx",
    "J": "berdasarkan jumlah.xlsx",
    "MASTER": "file barang.xlsx",
}
EXCLUDED_CODES = ["TC1777", "TC1778"]
INACTIVE_CODES = ["TC2134", "TH0339"]

# Baris sumber berikut bukan SKU stok mandiri lagi. Kode aslinya dipertahankan
# sebagai child virtual dan stoknya digabung ke induk dalam satuan dasar induk.
REPARENT_VARIANTS = [
    {"code": "TC1690", "parent_code": "TC0804", "factor": 0.4},
    {"code": "TC1765", "parent_code": "TC1756", "factor": 10.0},
    {"code": "TC1766", "parent_code": "TC1757", "factor": 10.0},
    {"code": "TC1767", "parent_code": "TC1758", "factor": 10.0},
    {"code": "TC1770", "parent_code": "TC1768", "factor": 0.23},
]

BOMS = [
    {"product_code": "1900", "materials": [{"code": "TC1898", "qty": 50.0}]},
    {"product_code": "TC1638", "materials": [{"code": "TC1089", "qty": 100.0}]},
    {"product_code": "TC1653", "materials": [{"code": "TC1652", "qty": 4.0}]},
    {"product_code": "TC1689", "materials": [{"code": "TC0804", "qty": 0.2}]},
    {"product_code": "TC1691", "materials": [{"code": "TC1637", "qty": 100.0}]},
    {"product_code": "TC1692", "materials": [{"code": "TC1637", "qty": 50.0}]},
    {"product_code": "TC1693", "materials": [{"code": "TC1087", "qty": 50.0}]},
    {
        "product_code": "TC1764",
        "materials": [
            {"code": "TC1763", "qty": 0.13},
            {"code": "TC1758", "qty": 1.0},
        ],
    },
]

COMMON_COLUMNS = [
    "KODEITEM", "NAMAITEM", "JENIS", "MEREK",
    "SATUAN1", "SATUAN2", "SATUAN3", "SATUAN4",
    "BARCODE1", "BARCODE2", "BARCODE3", "BARCODE4",
    "KONVERSI1", "KONVERSI2", "KONVERSI3", "KONVERSI4",
    "HARGAPOKOK1", "HARGAPOKOK2", "HARGAPOKOK3", "HARGAPOKOK4",
    "TIPE", "SUPPLIER", "SISTEMHPP", "KETERANGAN",
    "JENISPAJAK", "PAJAKINCLUDE",
]


def _text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


def _code(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    result = float(value)
    if not math.isfinite(result):
        raise ValueError(f"Nilai angka tidak valid: {value!r}")
    return result


def _json_number(value: float) -> int | float:
    value = round(float(value), 8)
    return int(value) if value.is_integer() else value


def _normalized_cell(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float):
        return _json_number(value)
    return value


def _load_sheet(path: Path, expected_mode: str) -> tuple[list[str], list[tuple[Any, ...]], dict[str, dict[str, Any]]]:
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Workbook contains no default style")
        workbook = load_workbook(path, read_only=False, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if len(rows) < 3 or _text(rows[0][0]) != "TIPE" or _text(rows[0][1]) != expected_mode:
        raise ValueError(f"Penanda mode workbook tidak sesuai: {path.name}")

    headers = [_text(value) for value in rows[1]]
    if not headers or headers[0] != "KODEITEM" or len(headers) != len(set(headers)):
        raise ValueError(f"Header workbook tidak valid: {path.name}")

    records: dict[str, dict[str, Any]] = {}
    normalized_rows: list[tuple[Any, ...]] = []
    for raw in rows[2:]:
        values = tuple(raw[: len(headers)])
        normalized_rows.append(tuple(_normalized_cell(value) for value in values))
        record = dict(zip(headers, values))
        code = _code(record.get("KODEITEM"))
        if not code:
            continue
        if code in records:
            raise ValueError(f"KODEITEM duplikat {code} pada {path.name}")
        records[code] = record
    return headers, normalized_rows, records


def _assert_sources_consistent(sheets: dict[str, tuple[list[str], list[tuple[Any, ...]], dict[str, dict[str, Any]]]]) -> None:
    j_headers, j_rows, j_records = sheets["J"]
    master_headers, master_rows, _ = sheets["MASTER"]
    if j_headers != master_headers or j_rows != master_rows:
        raise ValueError("file barang.xlsx harus identik dengan berdasarkan jumlah.xlsx")

    expected_codes = set(j_records)
    if len(expected_codes) != 1848:
        raise ValueError(f"Jumlah KODEITEM sumber berubah: {len(expected_codes)} (seharusnya 1848)")
    for mode in ("S", "L"):
        records = sheets[mode][2]
        if set(records) != expected_codes:
            missing = sorted(expected_codes - set(records))[:5]
            extra = sorted(set(records) - expected_codes)[:5]
            raise ValueError(f"Daftar kode mode {mode} berbeda; missing={missing}, extra={extra}")

    for code in sorted(expected_codes):
        baseline = j_records[code]
        for mode in ("S", "L"):
            candidate = sheets[mode][2][code]
            for column in COMMON_COLUMNS:
                left = _normalized_cell(baseline.get(column))
                right = _normalized_cell(candidate.get(column))
                if left != right:
                    raise ValueError(f"Master {code}.{column} berbeda antara J dan {mode}: {left!r} != {right!r}")


def _quantity_prices(record: dict[str, Any], slot: int) -> list[tuple[float, float]]:
    result = []
    for tier in range(1, 5):
        suffix = str(tier) if slot == 1 else f"{slot}_{tier}"
        qty = _number(record.get(f"JML{suffix}"))
        price = _number(record.get(f"HJ{suffix}"))
        if qty > 0 and price > 0:
            result.append((qty, price))
    return result


def _unit_price(slot: int, s_record: dict[str, Any], l_record: dict[str, Any], j_record: dict[str, Any]) -> dict[str, Any]:
    simple_price = _number(s_record.get(f"HARGAJUAL{slot}"))
    level_price = _number(l_record.get(f"HJ{slot}"))
    quantity_rows = _quantity_prices(j_record, slot)
    quantity_base = quantity_rows[0][1] if quantity_rows else 0.0

    # Level 1 adalah harga normal yang disepakati. Bila kosong, harga dasar mode
    # jumlah dipakai; mode satuan menjadi fallback terakhir.
    base_price = level_price or quantity_base or simple_price
    group_prices = {}
    for level in range(2, 5):
        price = _number(l_record.get(f"HJ{level}_{slot}"))
        if price > 0:
            group_prices[f"Level {level}"] = _json_number(price)

    tiers = []
    for index in range(1, len(quantity_rows)):
        previous_max = quantity_rows[index - 1][0]
        tiers.append({
            "min_qty": _json_number(previous_max + 1),
            "price": _json_number(quantity_rows[index][1]),
        })

    return {
        "sell_price": _json_number(base_price),
        "group_prices": group_prices,
        "tier_prices": tiers,
        "source_prices": {
            "S": _json_number(simple_price),
            "L": _json_number(level_price),
            "J": _json_number(quantity_base),
        },
    }


def build_payload(source_dir: Path) -> dict[str, Any]:
    sheets = {
        mode: _load_sheet(source_dir / filename, "J" if mode == "MASTER" else mode)
        for mode, filename in SOURCE_FILES.items()
    }
    _assert_sources_consistent(sheets)
    s_records = sheets["S"][2]
    l_records = sheets["L"][2]
    j_records = sheets["J"][2]

    items = []
    mode_counts = {"S": 0, "L": 0, "J": 0, "none": 0}
    for code in sorted(j_records):
        master = j_records[code]
        s_record = s_records[code]
        l_record = l_records[code]
        units = []
        active_modes = set()
        for slot in range(1, 5):
            unit_name = _text(master.get(f"SATUAN{slot}"))
            if not unit_name:
                continue
            pricing = _unit_price(slot, s_record, l_record, master)
            for mode, price in pricing["source_prices"].items():
                if price > 0:
                    active_modes.add(mode)
            units.append({
                "slot": slot,
                "name": unit_name,
                "conversion": _json_number(_number(master.get(f"KONVERSI{slot}")) or (1 if slot == 1 else 0)),
                "buy_price": _json_number(_number(master.get(f"HARGAPOKOK{slot}"))),
                "barcode": _code(master.get(f"BARCODE{slot}")) or None,
                **pricing,
            })
        if not units:
            raise ValueError(f"Barang {code} tidak memiliki satuan")
        if not any(abs(float(unit["conversion"]) - 1.0) < 1e-9 for unit in units):
            raise ValueError(f"Barang {code} tidak memiliki satuan KONVERSI=1")

        for mode in active_modes:
            mode_counts[mode] += 1
        if not active_modes:
            mode_counts["none"] += 1

        items.append({
            "code": code,
            "name": _text(master.get("NAMAITEM")),
            "category": _text(master.get("JENIS")) or "Umum",
            "brand": _text(master.get("MEREK")) or "Tanpa Merek",
            "supplier": _text(master.get("SUPPLIER")) or None,
            "stock": _json_number(_number(master.get("STOKAWAL"))),
            "min_stock": _json_number(_number(master.get("STOKMIN"))),
            "source_type": int(_number(master.get("TIPE")) or 1),
            "source_hpp_method": _text(master.get("SISTEMHPP")) or "FIFO",
            "hpp_method": "FIFO",
            "ppn_percent": _json_number(_number(master.get("PAJAKINCLUDE"))),
            "description": _text(master.get("KETERANGAN")) or None,
            "price_modes": sorted(active_modes),
            "units": units,
        })

    item_codes = {item["code"] for item in items}
    referenced = set(EXCLUDED_CODES) | set(INACTIVE_CODES)
    referenced.update(row["code"] for row in REPARENT_VARIANTS)
    referenced.update(row["parent_code"] for row in REPARENT_VARIANTS)
    for bom in BOMS:
        referenced.add(bom["product_code"])
        referenced.update(line["code"] for line in bom["materials"])
    missing = sorted(referenced - item_codes)
    if missing:
        raise ValueError(f"Kode koreksi tidak ditemukan pada sumber: {missing}")

    return {
        "schema_version": 1,
        "seed_version": SEED_VERSION,
        "source_files": SOURCE_FILES,
        "source_item_count": len(items),
        "price_mode_counts": mode_counts,
        "excluded_codes": EXCLUDED_CODES,
        "inactive_codes": INACTIVE_CODES,
        "reparent_variants": REPARENT_VARIANTS,
        "boms": BOMS,
        "items": items,
    }


def write_payload(payload: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    encoded = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    with output_path.open("wb") as raw:
        with gzip.GzipFile(filename="", mode="wb", fileobj=raw, mtime=0) as zipped:
            zipped.write(encoded)


def main() -> None:
    repo_root = Path(__file__).resolve().parent.parent
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-dir", type=Path, default=repo_root / "dataipos")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parent / "app" / "ipos_seed_v1.json.gz",
    )
    args = parser.parse_args()
    payload = build_payload(args.source_dir)
    write_payload(payload, args.output)
    print(
        f"Seed {payload['seed_version']} dibuat: {args.output} "
        f"({payload['source_item_count']} barang sumber)"
    )


if __name__ == "__main__":
    main()
